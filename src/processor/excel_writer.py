"""
Excel export for processed invoices.

Produces a professional multi-sheet workbook using pandas + openpyxl.
"""

from __future__ import annotations

from pathlib import Path
from typing import Sequence

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config as root_config
from src.processor.logger import get_logger
from src.processor.models import InvoiceData

logger = get_logger(__name__)


def invoices_to_dataframes(
    invoices: Sequence[InvoiceData],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Convert invoice models into header and line-item DataFrames."""
    headers = [inv.to_flat_dict() for inv in invoices]
    line_rows: list[dict] = []
    for inv in invoices:
        line_rows.extend(inv.line_items_as_rows())

    df_invoices = pd.DataFrame(headers)
    df_line_items = pd.DataFrame(line_rows)

    # TODO: Add CSV / JSON exporters that reuse these DataFrames.
    return df_invoices, df_line_items


def write_excel(
    invoices: Sequence[InvoiceData],
    output_path: str | Path,
    *,
    invoices_sheet: str | None = None,
    line_items_sheet: str | None = None,
) -> Path:
    """
    Write a professional Excel workbook with Invoices + LineItems sheets.

    Returns:
        Path to the written workbook.
    """
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    df_invoices, df_line_items = invoices_to_dataframes(invoices)
    sheet_invoices = invoices_sheet or root_config.EXCEL_SHEET_INVOICES
    sheet_lines = line_items_sheet or root_config.EXCEL_SHEET_LINE_ITEMS

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_invoices.to_excel(writer, sheet_name=sheet_invoices, index=False)
        df_line_items.to_excel(writer, sheet_name=sheet_lines, index=False)

        workbook = writer.book
        for sheet_name, dataframe in (
            (sheet_invoices, df_invoices),
            (sheet_lines, df_line_items),
        ):
            _style_worksheet(workbook[sheet_name], dataframe)

    logger.info("Wrote Excel workbook: %s (%s invoices)", output, len(invoices))
    return output


def _style_worksheet(worksheet, dataframe: pd.DataFrame) -> None:
    """Apply header styling and simple column autosizing."""
    header_fill = PatternFill("solid", fgColor="2C5282")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for idx, column in enumerate(dataframe.columns, start=1):
        if dataframe.empty:
            values: list[str] = []
        else:
            values = (
                dataframe.iloc[:, idx - 1]
                .fillna("")
                .astype(str)
                .replace("nan", "")
                .tolist()[:200]
            )
        max_len = max([len(str(column))] + [len(v) for v in values], default=10)
        worksheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 48)


# ---------------------------------------------------------------------------
# Future export backends
# ---------------------------------------------------------------------------

def write_csv(invoices: Sequence[InvoiceData], output_dir: str | Path) -> None:
    """TODO: Export invoices + line items as CSV files."""
    raise NotImplementedError("CSV export is on the roadmap.")


def write_json(invoices: Sequence[InvoiceData], output_path: str | Path) -> None:
    """TODO: Export invoices as a JSON document."""
    raise NotImplementedError("JSON export is on the roadmap.")
